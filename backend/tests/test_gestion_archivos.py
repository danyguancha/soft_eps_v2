"""
Tests para Gestión de Archivos (GA)
"""
import unittest
import os
from fastapi.testclient import TestClient
from io import BytesIO

from main import app

client = TestClient(app)

class TestGestionArchivos(unittest.TestCase):
    """Tests para endpoints de gestión de archivos"""
    
    @classmethod
    def setUpClass(cls):
        """Configuración inicial antes de todas las pruebas"""
        cls.test_files_dir = "test_files"
        os.makedirs(cls.test_files_dir, exist_ok=True)
        
        # Crear archivo CSV de prueba
        cls.csv_content = b"documento,nombre,fecha_nacimiento,departamento,municipio\n1234,Juan Perez,1990-01-01,CALDAS,MANIZALES"
        cls.csv_filename = "archivo1.csv"
        
        # Crear archivo Excel de prueba
        cls.excel_filename = "archivo2.xlsx"
        
        # Crear archivo TXT para prueba negativa
        cls.txt_content = b"Este es un archivo de texto plano"
        cls.txt_filename = "datos.txt"
    
    def test_GA_01(self):
        """GA-01: Subir archivo CSV válido y verificar que se procesa correctamente con todos sus metadatos"""
        # Crear archivo en memoria
        files = {
            'file': (self.csv_filename, BytesIO(self.csv_content), 'text/csv')
        }
        
        response = client.post("/api/v1/upload", files=files)
        
        # Validaciones
        self.assertEqual(response.status_code, 200)
        data = response.json()
        
        self.assertEqual(data["message"], "Archivo cargado exitosamente")
        self.assertIn("file_id", data)
        self.assertIsInstance(data["columns"], list)
        self.assertIn("documento", data["columns"])
        self.assertIn("nombre", data["columns"])
        self.assertEqual(data["sheets"], [])
        self.assertIsNone(data["default_sheet"])
        self.assertGreater(data["total_rows"], 0)
        self.assertEqual(data["is_excel"], False)
        self.assertEqual(data["has_sheets"], False)
        self.assertEqual(data["sheet_count"], 0)
        
        # sheet_detection_time puede ser None o 0.0 para CSV
        self.assertIn(data["sheet_detection_time"], [None, 0.0])
        
        self.assertEqual(data["ultra_fast"], True)
        
        # El engine puede tener diferentes valores
        self.assertIsInstance(data["engine"], str)
        self.assertTrue(
            "DuckDB" in data["engine"] or "Standard" in data["engine"],
            f"Engine inesperado: {data['engine']}"
        )
        self.assertGreaterEqual(data["file_size_mb"], 0)
        self.assertIsInstance(data["from_cache"], bool)
        
        print(f"GA-01 PASSED: {data['file_id']} cargado exitosamente")
    
    def test_GA_02(self):
        """GA-02: Subir archivo Excel con múltiples hojas y verificar detección correcta de todas las hojas"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Crear Excel con múltiples hojas
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Hoja1"
            ws1.append(["id", "nombre", "edad", "ciudad"])
            ws1.append([1, "Juan", 30, "Manizales"])
            ws2 = wb.create_sheet("Resumen")
            ws2.append(["total", "promedio"])
            ws2.append([100, 50])
            ws3 = wb.create_sheet("Datos2024")
            ws3.append(["fecha", "valor"])
            # Guardar en BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            files = {
                'file': (self.excel_filename, excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            response = client.post("/api/v1/upload", files=files)
            # Validaciones
            self.assertEqual(response.status_code, 200)
            data = response.json()
            
            self.assertEqual(data["message"], "Archivo cargado exitosamente")
            self.assertEqual(data["is_excel"], True)
            self.assertEqual(data["has_sheets"], True)
            self.assertEqual(data["sheet_count"], 3)
            self.assertIn("Hoja1", data["sheets"])
            self.assertIn("Resumen", data["sheets"])
            self.assertIn("Datos2024", data["sheets"])
            self.assertEqual(data["default_sheet"], "Hoja1")
            self.assertIsNotNone(data["sheet_detection_time"])
            
            print(f"GA-02 PASSED: Excel con {data['sheet_count']} hojas detectado")            
        except ImportError:
            self.skipTest("openpyxl no está instalado")
    
    def test_GA_03(self):
        """GA-03: Rechazar archivo con extensión no permitida (.txt)"""
        files = {
            'file': (self.txt_filename, BytesIO(self.txt_content), 'text/plain')
        }
        
        response = client.post("/api/v1/upload", files=files)
        
        # Validaciones
        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertIn("detail", data)
        
        # Validar el mensaje de error
        detail_lower = data["detail"].lower()
        self.assertTrue(
            "tipo de archivo no soportado" in detail_lower or 
            "extensión" in detail_lower or
            "no soportado" in detail_lower,
            f"Mensaje de error inesperado: {data['detail']}"
        )
        
        print(f"GA-03 PASSED: Archivo .txt correctamente rechazado")
        print(f"   Mensaje: {data['detail']}")
    
    @classmethod
    def tearDownClass(cls):
        """Limpieza después de todas las pruebas"""
        print("\n🧹 Limpieza completada")


if __name__ == "__main__":
    unittest.main(verbosity=2)
