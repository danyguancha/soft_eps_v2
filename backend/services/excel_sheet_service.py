# services/excel_sheet_service.py - NUEVO SERVICIO ULTRA-OPTIMIZADO
import os
import zipfile
import re
from typing import Dict, Any, Optional
from openpyxl import load_workbook
import pandas as pd
import time


class ExcelSheetService:
    """Servicio ultra-optimizado para obtener hojas de Excel sin cargar archivos completos"""
    
    @staticmethod
    def get_sheet_names_ultra_fast(file_path: str) -> Dict[str, Any]:
        """
        ULTRA-RÁPIDO: Obtiene nombres de hojas sin cargar el archivo completo
        Usa múltiples estrategias para archivos grandes
        """
        print(f"🔍 Obteniendo hojas de: {os.path.basename(file_path)}")
        start_time = time.time()
        
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": f"Archivo no encontrado: {file_path}",
                "sheets": [],
                "default_sheet": None
            }
        
        file_size_mb = os.path.getsize(file_path) / 1024 / 1024
        print(f"📊 Tamaño del archivo: {file_size_mb:.1f}MB")
        
        # Determinar estrategia según tamaño
        if file_size_mb > 50:  # Archivos > 50MB
            return ExcelSheetService._get_sheets_zipfile_method(file_path, start_time)
        elif file_size_mb > 20:  # Archivos > 20MB 
            return ExcelSheetService._get_sheets_openpyxl_readonly(file_path, start_time)
        else:  # Archivos < 20MB
            return ExcelSheetService._get_sheets_standard_method(file_path, start_time)
    
    @staticmethod
    def _get_sheets_zipfile_method(file_path: str, start_time: float) -> Dict[str, Any]:
        """ESTRATEGIA 1: Método ZIP ultra-rápido para archivos gigantes (>50MB)"""
        try:
            print("🚀 Estrategia ULTRA-RÁPIDA: Método ZIP directo")
            
            # Leer XML directamente del archivo ZIP
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                xml_content = zip_ref.read('docProps/app.xml').decode('utf-8')
            
            # Extraer nombres de hojas con regex
            titles_match = re.search(r'<TitlesOfParts>(.*?)</TitlesOfParts>', xml_content, re.DOTALL)
            
            if titles_match:
                titles_content = titles_match.group(1)
                # Extraer nombres individuales
                sheet_names = re.findall(r'<vt:lpstr>([^<]*)</vt:lpstr>', titles_content)
                
                if not sheet_names:
                    # Método alternativo si no encuentra lpstr
                    sheet_names = re.findall(r'>([^<>]+)<', titles_content)
                    sheet_names = [name.strip() for name in sheet_names if name.strip() and not name.isdigit()]
            else:
                sheet_names = []
            
            # Filtrar nombres válidos
            valid_sheets = [name for name in sheet_names if name and name.strip()]
            
            if not valid_sheets:
                # Fallback: usar método openpyxl
                print("⚠️ ZIP method no encontró hojas, usando fallback...")
                return ExcelSheetService._get_sheets_openpyxl_readonly(file_path, start_time)
            
            processing_time = time.time() - start_time
            
            print(f"Método ZIP completado en {processing_time:.3f}s:")
            print(f"   📋 {len(valid_sheets)} hojas encontradas: {valid_sheets}")
            
            return {
                "success": True,
                "sheets": valid_sheets,
                "default_sheet": valid_sheets[0] if valid_sheets else None,
                "method": "zipfile_ultra_fast",
                "processing_time": processing_time,
                "total_sheets": len(valid_sheets)
            }
            
        except Exception as e:
            print(f"❌ Método ZIP falló: {e}")
            # Fallback automático
            return ExcelSheetService._get_sheets_openpyxl_readonly(file_path, start_time)
    
    @staticmethod
    def _get_sheets_openpyxl_readonly(file_path: str, start_time: float) -> Dict[str, Any]:
        """ESTRATEGIA 2: OpenPyXL modo solo-lectura para archivos grandes (20-50MB)"""
        try:
            print("🔄 Estrategia RÁPIDA: OpenPyXL read-only")
            
            # Cargar en modo solo lectura (mucho más rápido)
            workbook = load_workbook(
                filename=file_path, 
                read_only=True,      # CRÍTICO: No carga datos, solo metadata
                keep_vba=False,      # No cargar macros
                data_only=True       # Solo valores, no fórmulas
            )
            
            sheet_names = workbook.sheetnames
            workbook.close()  # Liberar memoria inmediatamente
            
            processing_time = time.time() - start_time
            
            print(f"OpenPyXL read-only completado en {processing_time:.3f}s:")
            print(f"   📋 {len(sheet_names)} hojas: {sheet_names}")
            
            return {
                "success": True,
                "sheets": sheet_names,
                "default_sheet": sheet_names[0] if sheet_names else None,
                "method": "openpyxl_readonly",
                "processing_time": processing_time,
                "total_sheets": len(sheet_names)
            }
            
        except Exception as e:
            print(f"❌ OpenPyXL read-only falló: {e}")
            # Fallback final
            return ExcelSheetService._get_sheets_standard_method(file_path, start_time)
    
    @staticmethod
    def _get_sheets_standard_method(file_path: str, start_time: float) -> Dict[str, Any]:
        """ESTRATEGIA 3: Método estándar para archivos pequeños (<20MB)"""
        try:
            print("📄 Estrategia ESTÁNDAR: Pandas ExcelFile")
            
            # Usar pandas para archivos pequeños
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            sheet_names = excel_file.sheet_names
            excel_file.close()
            
            processing_time = time.time() - start_time
            
            print(f"Método estándar completado en {processing_time:.3f}s:")
            print(f"   📋 {len(sheet_names)} hojas: {sheet_names}")
            
            return {
                "success": True,
                "sheets": sheet_names,
                "default_sheet": sheet_names[0] if sheet_names else None,
                "method": "pandas_standard",
                "processing_time": processing_time,
                "total_sheets": len(sheet_names)
            }
            
        except Exception as e:
            print(f"❌ Método estándar falló: {e}")
            
            # Último recurso: devolver hoja por defecto
            return {
                "success": False,
                "error": f"No se pudieron obtener las hojas: {str(e)}",
                "sheets": ["Sheet1"],  # Hoja por defecto como fallback
                "default_sheet": "Sheet1",
                "method": "fallback_default",
                "processing_time": time.time() - start_time,
                "total_sheets": 1
            }

    @staticmethod
    def validate_sheet_exists(file_path: str, sheet_name: str) -> bool:
        """Valida que una hoja específica existe en el archivo"""
        try:
            sheet_info = ExcelSheetService.get_sheet_names_ultra_fast(file_path)
            if sheet_info["success"]:
                return sheet_name in sheet_info["sheets"]
            return False
        except:
            return False

    @staticmethod
    def get_first_valid_sheet(file_path: str) -> Optional[str]:
        """Obtiene la primera hoja válida del archivo"""
        try:
            sheet_info = ExcelSheetService.get_sheet_names_ultra_fast(file_path)
            if sheet_info["success"] and sheet_info["sheets"]:
                return sheet_info["sheets"][0]
            return None
        except:
            return None


# Instancia global del servicio
excel_sheet_service = ExcelSheetService()
