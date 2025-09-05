# services/excel_service.py
import pandas as pd
from openpyxl import load_workbook
from typing import Iterator, Optional, List, Dict, Any
from pathlib import Path
from models.base import AbstractFile
from services.interfaces import IFileService


def convert_excel_dates_to_readable(series):
    """Convierte fechas tipo Excel (números) y strings a formato DD/MM/YYYY"""
    result = []
    for val in series:
        # Si es nulo o string vacío
        if pd.isna(val) or str(val).lower() in ['nan', 'null', '', 'none']:
            result.append('')
            continue
        
        # Si es un número (fecha serial de Excel)
        try:
            as_float = float(val)
            # Solo convertir si parece ser una fecha serial válida (> 2000 para años recientes)
            if as_float > 2000:
                d = pd.to_datetime(as_float, unit='D', origin='1899-12-30')
                result.append(d.strftime('%d/%m/%Y'))
                continue
        except (ValueError, TypeError):
            pass
        
        # Si es una string que ya parece fecha
        try:
            d = pd.to_datetime(str(val), errors='coerce', dayfirst=True)
            if pd.notna(d):
                result.append(d.strftime('%d/%m/%Y'))
            else:
                result.append(str(val))
        except Exception:
            result.append(str(val))
    
    return result


class ExcelFile(AbstractFile):
    def __init__(self, file_path: str):
        super().__init__(file_path)
        self._sheets_cache: List[str] | None = None
        self._columns_cache: Dict[str, List[str]] = {}
        self._total_rows_cache: Dict[str, int] = {}

    def get_sheets(self) -> List[str]:
        """Obtiene las hojas del archivo, con cache"""
        if self._sheets_cache is not None:
            return self._sheets_cache
            
        try:
            wb = load_workbook(self.file_path, read_only=True)
            self._sheets_cache = wb.sheetnames
            wb.close()
            return self._sheets_cache
        except Exception as e:
            print(f"Error obteniendo hojas: {e}")
            return []

    def count_rows_efficiently(self, sheet_name: str) -> int:
        """Cuenta filas de manera eficiente usando openpyxl"""
        if sheet_name in self._total_rows_cache:
            return self._total_rows_cache[sheet_name]
            
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            total_rows = ws.max_row - 1 if ws.max_row else 0
            wb.close()
            self._total_rows_cache[sheet_name] = max(0, total_rows)
            return self._total_rows_cache[sheet_name]
        except Exception as e:
            print(f"Error contando filas: {e}")
            try:
                df = pd.read_excel(
                    self.file_path, 
                    sheet_name=sheet_name, 
                    usecols=[0],
                    engine='openpyxl'
                )
                count = len(df)
                self._total_rows_cache[sheet_name] = count
                return count
            except Exception:
                return 0

    def _detect_date_columns(self, columns: List[str]) -> List[str]:
        """Detecta columnas que probablemente contienen fechas"""
        date_indicators = [
            'fecha', 'date', 'datetime', 'time', 'nacimiento', 'prestacion', 
            'radicacion', 'conciliacion', 'retiro', 'afiliacion', 'fecnac'
        ]
        
        date_columns = []
        for col in columns:
            col_lower = col.lower()
            if any(indicator in col_lower for indicator in date_indicators):
                date_columns.append(col)
        
        return date_columns

    def _convert_date_columns(self, df: pd.DataFrame, date_columns: List[str]) -> pd.DataFrame:
        """Convierte columnas detectadas a formato fecha legible"""
        for col in date_columns:
            if col in df.columns:
                try:
                    df[col] = convert_excel_dates_to_readable(df[col])
                    print(f"Columna '{col}' convertida a fecha")
                except Exception as e:
                    print(f"⚠️ No se pudo convertir '{col}' a fecha: {e}")
        
        return df

    def get_data(self, sheet_name: str = None) -> pd.DataFrame:
        """Cargar Excel con conversión automática de fechas"""
        target_sheet = sheet_name or self.get_sheets()[0] if self.get_sheets() else 0
        
        # Leer primero como string para poder convertir fechas
        df = pd.read_excel(
            self.file_path, 
            sheet_name=target_sheet,
            engine='openpyxl',
            dtype=str
        )
        
        # Detectar y convertir columnas de fecha
        date_columns = self._detect_date_columns(df.columns.tolist())
        df = self._convert_date_columns(df, date_columns)
        
        return df

    def get_columns(self, sheet_name: str) -> List[str]:
        """Obtiene columnas de manera eficiente leyendo solo el header"""
        if sheet_name in self._columns_cache:
            return self._columns_cache[sheet_name]
            
        try:
            df_header = pd.read_excel(
                self.file_path, 
                sheet_name=sheet_name, 
                nrows=0,
                engine='openpyxl'
            )
            columns = df_header.columns.tolist()
            self._columns_cache[sheet_name] = columns
            return columns
        except Exception as e:
            print(f"Error obteniendo columnas: {e}")
            return []

    def get_data_chunked(
        self, 
        sheet_name: str,
        start_row: int = 0, 
        nrows: Optional[int] = None
    ) -> pd.DataFrame:
        """Obtiene datos por chunks con conversión de fechas"""
        read_params = {
            'sheet_name': sheet_name,
            'engine': 'openpyxl',
            'dtype': str
        }
        
        if start_row > 0:
            read_params['skiprows'] = start_row
        
        if nrows is not None:
            read_params['nrows'] = nrows
        
        df = pd.read_excel(self.file_path, **read_params)
        
        # Convertir fechas
        date_columns = self._detect_date_columns(df.columns.tolist())
        df = self._convert_date_columns(df, date_columns)
        
        return df

    def get_chunk_iterator(
        self, 
        sheet_name: str, 
        chunk_size: int = 10000
    ) -> Iterator[pd.DataFrame]:
        """Retorna un iterador para procesar la hoja por chunks con conversión de fechas"""
        total_rows = self.count_rows_efficiently(sheet_name)
        
        for start_row in range(0, total_rows, chunk_size):
            rows_to_read = min(chunk_size, total_rows - start_row)
            
            if rows_to_read <= 0:
                break
                
            yield self.get_data_chunked(
                sheet_name=sheet_name,
                start_row=start_row,
                nrows=rows_to_read
            )

    def get_sample_data(self, sheet_name: str, n_rows: int = 1000) -> pd.DataFrame:
        """Obtiene una muestra pequeña para inspección rápida"""
        return self.get_data_chunked(sheet_name, start_row=0, nrows=n_rows)

    def get_file_info(self) -> Dict[str, Any]:
        """Retorna información del archivo sin cargarlo completo"""
        sheets = self.get_sheets()
        info = {
            'sheets': sheets,
            'has_sheets': True,
            'file_size_bytes': Path(self.file_path).stat().st_size,
            'sheets_info': {}
        }
        
        for sheet in sheets:
            info['sheets_info'][sheet] = {
                'total_rows': self.count_rows_efficiently(sheet),
                'columns': self.get_columns(sheet)
            }
        
        return info


class ExcelService(IFileService):
    def load(self, file_path: str) -> ExcelFile:
        if not Path(file_path).is_file():
            raise FileNotFoundError(file_path)
        return ExcelFile(file_path)
    
    def get_columns(self, obj: ExcelFile, sheet_name: str) -> list:
        return obj.get_columns(sheet_name)
    
    def get_sheets(self, obj: ExcelFile) -> list:
        return obj.get_sheets()
    
    def get_data(self, obj: ExcelFile, sheet_name: str = None) -> pd.DataFrame:
        return obj.get_data(sheet_name)
    
    def get_data_chunked(
        self, 
        obj: ExcelFile, 
        sheet_name: str,
        start_row: int = 0, 
        nrows: Optional[int] = None
    ) -> pd.DataFrame:
        return obj.get_data_chunked(sheet_name, start_row, nrows)
    
    def get_chunk_iterator(
        self, 
        obj: ExcelFile, 
        sheet_name: str, 
        chunk_size: int = 10000
    ) -> Iterator[pd.DataFrame]:
        return obj.get_chunk_iterator(sheet_name, chunk_size)
    
    def get_file_info(self, obj: ExcelFile) -> Dict[str, Any]:
        return obj.get_file_info()
