import os
import time
import zipfile
import re
import pandas as pd
from typing import Dict, Any

class ExcelSheetsController:
    """Controlador para manejo de hojas Excel"""
    
    def get_excel_sheets(self, file_path: str) -> Dict[str, Any]:
        """Obtiene hojas de Excel de forma ultra-optimizada"""
        start_time = time.time()
        
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": f"Archivo no encontrado: {file_path}",
                "sheets": [],
                "default_sheet": None
            }
        
        file_size_mb = os.path.getsize(file_path) / 1024 / 1024
        
        # Determinar estrategia según tamaño
        if file_size_mb > 50:  # Archivos > 50MB
            return self._get_sheets_zipfile_method(file_path, start_time)
        elif file_size_mb > 20:  # Archivos > 20MB 
            return self._get_sheets_openpyxl_readonly(file_path, start_time)
        else:  # Archivos < 20MB
            return self._get_sheets_standard_method(file_path, start_time)

    def _get_sheets_zipfile_method(self, file_path: str, start_time: float) -> Dict[str, Any]:
        """Método ZIP ultra-rápido para archivos gigantes (>50MB)"""
        try:            
            # Leer XML directamente del archivo ZIP
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                xml_content = zip_ref.read('docProps/app.xml').decode('utf-8')
            
            # Extraer nombres de hojas con regex
            titles_match = re.search(r'<TitlesOfParts>(.*?)</TitlesOfParts>', xml_content, re.DOTALL)
            
            if titles_match:
                titles_content = titles_match.group(1)
                # Extraer nombres individuales
                sheet_names = re.findall(r'<vt:lpstr>([^<]*)</vt:lpstr>', titles_content)
                
                if not sheet_names:
                    # Método alternativo si no encuentra lpstr
                    sheet_names = re.findall(r'>([^<>]+)<', titles_content)
                    sheet_names = [name.strip() for name in sheet_names if name.strip() and not name.isdigit()]
            else:
                sheet_names = []
            
            # Filtrar nombres válidos
            valid_sheets = [name for name in sheet_names if name and name.strip()]
            
            if not valid_sheets:
                # Fallback: usar método openpyxl
                return self._get_sheets_openpyxl_readonly(file_path, start_time)
            
            processing_time = time.time() - start_time
            
            return {
                "success": True,
                "sheets": valid_sheets,
                "default_sheet": valid_sheets[0] if valid_sheets else None,
                "method": "zipfile_ultra_fast",
                "processing_time": processing_time,
                "total_sheets": len(valid_sheets)
            }
            
        except Exception as e:
            print(f"Método ZIP falló: {e}")
            # Fallback automático
            return self._get_sheets_openpyxl_readonly(file_path, start_time)

    def _get_sheets_openpyxl_readonly(self, file_path: str, start_time: float) -> Dict[str, Any]:
        """OpenPyXL modo solo-lectura para archivos grandes (20-50MB)"""
        try:            
            from openpyxl import load_workbook
            
            # Cargar en modo solo lectura (mucho más rápido)
            workbook = load_workbook(
                filename=file_path, 
                read_only=True,      # CRÍTICO: No carga datos, solo metadata
                keep_vba=False,      # No cargar macros
                data_only=True       # Solo valores, no fórmulas
            )
            
            sheet_names = workbook.sheetnames
            workbook.close()  # Liberar memoria inmediatamente
            
            processing_time = time.time() - start_time
            
            return {
                "success": True,
                "sheets": sheet_names,
                "default_sheet": sheet_names[0] if sheet_names else None,
                "method": "openpyxl_readonly",
                "processing_time": processing_time,
                "total_sheets": len(sheet_names)
            }
            
        except Exception:
            return self._get_sheets_standard_method(file_path, start_time)

    def _get_sheets_standard_method(self, file_path: str, start_time: float) -> Dict[str, Any]:
        """Método estándar para archivos pequeños (<20MB)"""
        try:            
            # Usar pandas para archivos pequeños
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            sheet_names = excel_file.sheet_names
            excel_file.close()
            
            processing_time = time.time() - start_time            
            return {
                "success": True,
                "sheets": sheet_names,
                "default_sheet": sheet_names[0] if sheet_names else None,
                "method": "pandas_standard",
                "processing_time": processing_time,
                "total_sheets": len(sheet_names)
            }
            
        except Exception as e:            
            # Último recurso: devolver hoja por defecto
            return {
                "success": False,
                "error": f"No se pudieron obtener las hojas: {str(e)}",
                "sheets": ["Sheet1"], 
                "default_sheet": "Sheet1",
                "method": "fallback_default",
                "processing_time": time.time() - start_time,
                "total_sheets": 1
            }

    def validate_sheet_exists(self, file_path: str, sheet_name: str) -> bool:
        """Valida que una hoja específica existe en el archivo"""
        try:
            sheet_info = self.get_excel_sheets(file_path)
            if sheet_info["success"]:
                return sheet_name in sheet_info["sheets"]
            return False
        except Exception:
            return False

    def get_first_valid_sheet(self, file_path: str) -> str:
        """Obtiene la primera hoja válida del archivo"""
        try:
            sheet_info = self.get_excel_sheets(file_path)
            if sheet_info["success"] and sheet_info["sheets"]:
                return sheet_info["sheets"][0]
            return ""
        except Exception:
            return ""
